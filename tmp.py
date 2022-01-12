import logging
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import column_index_from_string
import pprint
import re


def compareXl(wb1 : openpyxl.Workbook, wb2 : openpyxl.Workbook):
    logging.info('Function {} starts'.format('compareXl'))
    
    wb1data = analyzeWorkbook(wb1)
    wb2data = analyzeWorkbook(wb2)

    changedCells = []
    formulaCells = []
    for sheetname in wb1data.keys():
        wb1cells = wb1data[sheetname]
        wb2cells = wb2data[sheetname]
        for coordinate in wb1cells.keys():
            wb1val = wb1[sheetname][coordinate].value
            if str(wb1val).startswith('='):
                formulaCells.append({'sheet' :  sheetname, 'cell' : coordinate, 'value' : wb1val})

            if wb1cells[coordinate] != wb2cells[coordinate]:
                changedCells.append({'sheet' :  sheetname, 'cell' : coordinate})

    logging.info('changed cells')
    logging.info(pprint.pformat(changedCells))
    logging.info('formulas')
    logging.info(pprint.pformat(formulaCells))

    # find cells referencing changedCells recursively
    changedCells = findRefs(changedCells,formulaCells)
    # Highlight Changes
    for change in changedCells:
        changedSheet = change['sheet']
        chengedCordinate = change['cell']
        fill = PatternFill(patternType='solid', fgColor='00FF00')
        wb1[changedSheet][chengedCordinate].fill = fill

    logging.info('Function {} ends'.format('compareXl'))
    return wb1

def findRefs(changes, formulas):
    logging.info('Function {} starts'.format('findRefs'))
    logging.info('{} changes'.format(len(changes)))
    logging.info('{} formulas'.format(len(formulas)))

    # Check changed cell is in formulacell
    # if found, move formula cell into changes and call this func recursively
    # If no changes are referenced by formula, this func returns changes
    for change in changes:
        for formula in formulas:
            if isReferenced(change,formula):
                formulas.remove(formula)
                changes.append(formula)
                findRefs(changes,formulas)
    logging.info('Function {} ends'.format('findRefs'))
    return changes

def isReferenced(change, formula):
    changeSheet = change["sheet"]
    changeCellCordinate = change["cell"]
    formulaSheet = formula["sheet"]
    formulaexpression = formula["value"]

    logging.info("changeSheet {}".format(changeSheet))
    logging.info("changeCellCordinate {}".format(changeCellCordinate))
    logging.info("formulaSheet {}".format(formulaSheet))
    logging.info("formulaexpression {}".format(formulaexpression))

    # Tokenize formula and extract RANGE
    tok = Tokenizer(formulaexpression)
    #print("\n".join("%12s%11s%9s" % (t.value, t.type, t.subtype) for t in tok.items))
    for token in tok.items:
        if token.subtype == "RANGE":
            logging.info("RANGE Found {}".format(token.value))
            range = token.value
            targetSheetName = formulaSheet
            if "!" in range:
                targetSheetName = range[:range.find('!')]
                targetCordinate = range[range.find('!') +1:]
            else:
                targetCordinate = range
            
            if targetSheetName != changeSheet:
                continue

            logging.info("targetSheetName {}".format(targetSheetName))
            logging.info("targetCordinate {}".format(targetCordinate))
            logging.info("changeCellCordinate {}".format(changeCellCordinate))


            # Check changed cell cordinate is in the range or not
            targetCordinate = targetCordinate.replace("$","")
            changeCellCordinate = changeCellCordinate.replace("$","")

            if ":" not in targetCordinate:
                targetCordinate = "{}:{}".format(targetCordinate,targetCordinate)
            
            if ":" not in changeCellCordinate:
                changeCellCordinate = "{}:{}".format(changeCellCordinate,changeCellCordinate)

            m = re.search("(.+):(.+)",targetCordinate)
            targetStartCordinate = m.group(1)
            targetEndCordinate = m.group(2)
            logging.info('targetStartCordinate {}'.format(targetStartCordinate))
            logging.info('targetEndCordinate {}'.format(targetEndCordinate))

            m = re.search("([A-Z]+)([0-9]+)",targetStartCordinate)
            targetStartCol = int(column_index_from_string(m.group(1)))
            targetStartRow = int(m.group(2))
            logging.info('targetStartCol {}'.format(targetStartCol))
            logging.info('targetStartRow {}'.format(targetStartRow))

            m = re.search("([A-Z]+)([0-9]+)",targetEndCordinate)
            targetEndCol = int(column_index_from_string(m.group(1)))
            targetEndRow = int(m.group(2))
            logging.info('targetEndCol {}'.format(targetEndCol))
            logging.info('targetStartRow {}'.format(targetEndRow))

            
            m = re.search("(.+):(.+)",changeCellCordinate)
            changeStartCordinate = m.group(1)
            changeEndCordinate = m.group(2)
            logging.info('changeStartCordinate {}'.format(changeStartCordinate))
            logging.info('changeEndCordinate {}'.format(changeEndCordinate))

            m = re.search("([A-Z]+)([0-9]+)",changeStartCordinate)
            changeStartCol = int(column_index_from_string(m.group(1)))
            changeStartRow = int(m.group(2))
            logging.info('changeStartCol {}'.format(changeStartCol))
            logging.info('changeStartRow {}'.format(changeStartRow))

            m = re.search("([A-Z]+)([0-9]+)",changeEndCordinate)
            changeEndCol = int(column_index_from_string(m.group(1)))
            changeEndRow = int(m.group(2))
            logging.info('changeEndCol {}'.format(changeEndCol))
            logging.info('changeEndRow {}'.format(changeEndRow))

            # Check change start cell is in the target range
            if changeStartRow >= targetStartRow and changeStartRow <= targetEndRow and changeStartCol >= targetStartCol and changeStartCol <= targetEndCol:
                logging.info('isReferenced: {}'.format("True"))
                return True

            # Check change end cell is in the target range
            if changeEndRow >= targetStartRow and changeEndRow <= targetEndRow and changeEndCol >= targetStartCol and changeEndCol <= targetEndCol:
                logging.info('isReferenced: {}'.format("True"))
                return True

    logging.info('isReferenced: {}'.format("False"))
    return False

def analyzeWorkbook(wb : openpyxl.Workbook):
    result = {}
    for sheet in wb:
        sheetname = sheet.title
        cells = {}
        for row in sheet:
            for cell in row:
                cells[cell.coordinate] = cell.value
        result[sheetname] = cells
    
    logging.info(pprint.pformat(result))
    return result

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)

    excelfilepath1 = "./xlsx_files/file2.xlsx"
    excelfilepath2 = "./xlsx_files/file1.xlsx"

    wb1 = openpyxl.load_workbook(excelfilepath1, data_only=False, keep_vba=False)
    wb2 = openpyxl.load_workbook(excelfilepath2, data_only=False, keep_vba=False)

    wbComp = compareXl(wb1,wb2)
    wbComp.save("./xlsx_files/compare.xlsx")
    wb1.close()
    wb2.close()
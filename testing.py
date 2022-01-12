from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

# Reading a XLSX file
wb = load_workbook(filename="./xlsx_files/file1.xlsx")

# Printing sheets names in a excel (XLSX) file.
sheet_names = wb.sheetnames
print('Printing all sheet names : ', sheet_names)

# Printing 
sheet = wb.active
print('Printing which one of the sheet is active: ', sheet.title)

for row in sheet.iter_rows(min_row=1, min_col=1,values_only=True):
    print(row)
    

# for col in sheet.iter_cols(min_row=1, min_col=1, values_only=True):
#     print(col)


# for row in sheet.rows:
#     print(row)xds